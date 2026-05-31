import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def convert_to_excel():
    """Convert test_cases.json to readable Excel format with merged cells"""
    
    # Load test cases from JSON
    print("Loading test_cases.json...")
    with open('test_cases.json', 'r', encoding='utf-8') as f:
        test_cases = json.load(f)
    
    print(f"Loaded {len(test_cases)} test cases")
    
    # Create new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    
    # Define column headers
    headers = [
        "TC_ID",
        "Scenario",
        "Description",
        "Steps",
        "Expected Result",
        "Actual Result",
        "Status"
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Write test case data with merged cells
    print("Writing test cases to Excel...")
    current_row = 2
    
    for tc in test_cases:
        tc_id = tc['TC_ID']
        scenario = tc['Scenario']
        description = tc['Description']
        steps = tc['steps']
        
        # Calculate the range for this test case (number of steps)
        num_steps = len(steps)
        start_row = current_row
        end_row = current_row + num_steps - 1
        
        # Write TC_ID (merged across all steps)
        ws.cell(row=start_row, column=1, value=tc_id)
        if num_steps > 1:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        
        # Write Scenario (merged across all steps)
        ws.cell(row=start_row, column=2, value=scenario)
        if num_steps > 1:
            ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
        
        # Write Description (merged across all steps)
        ws.cell(row=start_row, column=3, value=description)
        if num_steps > 1:
            ws.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)
        
        # Write each step on its own row
        for step_idx, step in enumerate(steps):
            row = current_row + step_idx
            ws.cell(row=row, column=4, value=step['action'])
            ws.cell(row=row, column=5, value=step['expected'])
            ws.cell(row=row, column=6, value="")  # Actual Result (empty)
            ws.cell(row=row, column=7, value="")  # Status (empty)
        
        # Color code based on scenario (module)
        if scenario == "Login":
            fill_color = "FFC7CE"  # Light red
        elif scenario == "Cart":
            fill_color = "FFE699"  # Light yellow
        elif scenario == "Address":
            fill_color = "C6E0B4"  # Light green
        elif scenario == "Delivery":
            fill_color = "D9EAD3"  # Light green-blue
        elif scenario == "Payment":
            fill_color = "E1BEE7"  # Light purple
        elif scenario == "Order Review":
            fill_color = "FFCCBC"  # Light orange
        elif scenario == "Order Confirmation":
            fill_color = "B2DFDB"  # Light teal
        elif scenario == "Edge Cases":
            fill_color = "FFF9C4"  # Light yellow
        elif scenario == "Negative Cases":
            fill_color = "FFCDD2"  # Light pink
        elif scenario == "Performance":
            fill_color = "D1C4E9"  # Light lavender
        else:
            fill_color = None
        
        if fill_color:
            for col_num in range(1, 8):
                for row in range(start_row, end_row + 1):
                    cell = ws.cell(row=row, column=col_num)
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        # Wrap text for better readability
        for col_num in range(1, 8):
            for row in range(start_row, end_row + 1):
                cell = ws.cell(row=row, column=col_num)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        current_row += num_steps
    
    # Set column widths
    print("Setting column widths...")
    column_widths = {
        1: 12,   # TC_ID
        2: 15,   # Scenario
        3: 50,   # Description
        4: 40,   # Steps
        5: 50,   # Expected Result
        6: 30,   # Actual Result
        7: 15    # Status
    }
    
    for col_num, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save workbook
    output_file = "test_cases_readable_new.xlsx"
    print(f"Saving to {output_file}...")
    wb.save(output_file)
    print(f"Successfully created {output_file}")
    print(f"Total test cases: {len(test_cases)}")
    print(f"Total rows: {current_row - 1}")

if __name__ == "__main__":
    convert_to_excel()
